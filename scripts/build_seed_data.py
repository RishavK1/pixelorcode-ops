from pathlib import Path
import json

from openpyxl import load_workbook


ROOT = Path("/Users/rishav/Downloads/codex")
OUT = ROOT / "pixelorcode-ops" / "src" / "seedData.json"

SOURCES = [
    ("Bangalore Clinics", ROOT / "outputs/bangalore_clinic_leads/bangalore_no_website_clinic_leads.xlsx", "Clinic"),
    ("Bangalore Education", ROOT / "outputs/bangalore_education_leads/bangalore_no_website_education_leads.xlsx", "Education"),
    ("Pune Clinics", ROOT / "outputs/pune_clinic_leads/pune_no_website_clinic_leads.xlsx", "Clinic"),
    ("Pune Education", ROOT / "outputs/pune_education_leads/pune_no_website_education_leads.xlsx", "Education"),
    ("UAE Clinics", ROOT / "outputs/uae_clinic_leads/uae_dental_aesthetic_clinic_leads.xlsx", "Clinic"),
    ("International Service Businesses", ROOT / "outputs/international_service_business_leads/international_normal_service_business_leads_400.xlsx", "Service Business"),
    ("International Startups", ROOT / "outputs/international_startup_leads/international_non_ai_small_startup_leads_250.xlsx", "Startup"),
]


def pick(row, headers, names):
    for name in names:
        if name in headers:
            value = row[headers[name]]
            if value is not None and str(value).strip():
                return str(value).strip()
    return ""


def infer_status(row, headers):
    raw = pick(row, headers, ["Call Status", "Email Status", "Outreach Status", "Status"])
    if not raw:
        return "Not Contacted"
    value = raw.lower()
    if value in {"not called", "not contacted", "new"}:
        return "Not Contacted"
    if "bounce" in value:
        return "Bounced"
    if "done" in value or "sent" in value:
        return "Email Sent"
    if "interested" in value:
        return "Interested"
    if "follow" in value:
        return "Follow Up"
    if "wrong" in value:
        return "Wrong Number"
    return raw


def is_bangalore_education_sent(row_number):
    first_batch_sent = {1, 2, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20}
    second_batch_sent = {21, 22, 23, 25, 26, 27, 28, 30, 31, 32, 33, 35, 36, 37, 38, 39, 40}
    final_batch_sent = set(range(41, 54))
    return row_number in first_batch_sent or row_number in second_batch_sent or row_number in final_batch_sent


def read_sheet(list_name, path, fallback_niche):
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    headers = {name: idx for idx, name in enumerate(header_row) if name}
    leads = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not any(values):
            continue
        name = pick(values, headers, ["Business Name", "Clinic Name", "Company Name", "Startup Name", "Lead Name", "Name"])
        if not name:
            continue
        source_id = pick(values, headers, ["Lead ID", "ID"]) or f"{list_name[:3].upper()}-{idx:03d}"
        status = infer_status(values, headers)
        website_status = pick(values, headers, ["Website Status", "Website Quality", "Website"])
        email_sent = status in {"Done", "Bounced", "Email Sent"} or "sent" in str(status).lower()
        whatsapp_sent = list_name == "Bangalore Education" and is_bangalore_education_sent(idx)
        lead = {
            "id": f"{list_name.lower().replace(' ', '-')}-{idx:03d}",
            "sourceId": source_id,
            "name": name,
            "list": list_name,
            "niche": pick(values, headers, ["Specialty", "Education Segment", "Industry", "Business Type", "Niche"]) or fallback_niche,
            "location": pick(values, headers, ["Area / Locality", "Location", "City", "Country"]),
            "address": pick(values, headers, ["Full Address", "Address"]),
            "phone": pick(values, headers, ["Phone Number", "Phone", "Primary Phone"]),
            "alternatePhone": pick(values, headers, ["Alternate Phone", "Alt Phone"]),
            "email": pick(values, headers, ["Email", "Email Address"]),
            "websiteStatus": website_status or "Unknown",
            "rating": pick(values, headers, ["Rating"]),
            "reviews": pick(values, headers, ["Review Count", "Reviews"]),
            "sourceLink": pick(values, headers, ["Google Maps / Directory Link", "Primary Source Link", "Source Link", "Website", "Company Website"]),
            "otherLink": pick(values, headers, ["Other Profile Link", "LinkedIn", "LinkedIn URL"]),
            "leadReason": pick(values, headers, ["Why This Is A Good Lead", "Lead Description", "Why Good Lead"]),
            "pitch": pick(values, headers, ["Suggested Call Pitch", "Pitch", "Suggested Email Angle"]),
            "decisionMaker": pick(values, headers, ["Decision Maker / Contact Person", "Founder Name", "Contact Person"]),
            "openingHours": pick(values, headers, ["Opening Hours"]),
            "notes": pick(values, headers, ["Notes", "Source Verification Notes", "Website Verification Notes"]),
            "status": status,
            "owner": "Unassigned",
            "lastAction": status if status not in {"New", "Not Contacted"} else "Imported",
            "whatsappSent": whatsapp_sent,
            "whatsappReplied": False,
            "emailSent": email_sent,
            "nextFollowUp": "",
            "proposalStatus": "None",
            "clientValue": "",
            "createdAt": "2026-06-05",
            "updatedAt": "2026-06-05",
        }
        leads.append(lead)
    return leads


def apply_outreach_updates(leads):
    by_name = {lead["name"].strip().lower(): lead for lead in leads}

    def update(name, **fields):
        lead = by_name.get(name.lower())
        if lead:
            lead.update(fields)
            lead["updatedAt"] = "2026-06-05"

    update(
        "Hebbar's Academy",
        status="Replied",
        owner="Sales Team",
        whatsappSent=True,
        whatsappReplied=True,
        lastAction="WhatsApp reply received: asked to share details",
        nextFollowUp="2026-06-05",
        notes="WhatsApp outreach response received. Send concise website/app/automation details and ask for a short call.",
    )
    update(
        "Atrian Tuition Centre",
        status="Replied",
        owner="Sales Team",
        whatsappSent=True,
        whatsappReplied=True,
        lastAction="WhatsApp reply received: interested in details",
        nextFollowUp="2026-06-05",
        notes="WhatsApp outreach response received. Good follow-up candidate for website plus enquiry automation pitch.",
    )
    update(
        "Ryan's Tuition Centre",
        status="Proposal Sent",
        owner="Sales Team",
        whatsappSent=True,
        whatsappReplied=True,
        lastAction="Proposal sent; project agreed verbally; design work in progress",
        nextFollowUp="2026-06-06",
        proposalStatus="Sent",
        clientValue="26000",
        notes="Not marked closed until payment is received. Proposal sent for website + small CRM; design due Saturday.",
    )

    for lead in leads:
        if lead["list"] == "Bangalore Education" and lead.get("whatsappSent") and not lead.get("whatsappReplied"):
            lead["status"] = "WhatsApp Sent"
            lead["owner"] = "Sales Team"
            lead["lastAction"] = "WhatsApp outreach sent"

    return leads


def main():
    leads = []
    for list_name, path, niche in SOURCES:
        leads.extend(read_sheet(list_name, path, niche))
    leads = apply_outreach_updates(leads)

    payload = {
        "__version": "2026-06-05-not-contacted-v3",
        "generatedAt": "2026-06-05",
        "leads": leads,
        "proposals": [
            {
                "id": "proposal-ryan-001",
                "leadName": "Ryan's Tuition Centre",
                "client": "Ryan's Tuition Centre",
                "status": "Sent",
                "value": "26000",
                "phone": "78923 01731",
                "email": "Not found publicly",
                "owner": "Sales Team",
                "service": "Website + small CRM",
                "sentDate": "2026-06-04",
                "validUntil": "2026-07-04",
                "nextStep": "Design in progress; payment pending before marking closed",
                "file": "/Users/rishav/Documents/Proposal - Pixelorcode x Ryan tution center.pdf",
                "fileUrl": "/proposals/ryan-tuition-centre-proposal.pdf",
                "fileName": "Proposal - Pixelorcode x Ryan tution center.pdf",
                "fileType": "application/pdf",
                "notes": "Website + small CRM proposal sent. Do not mark closed until payment is received."
            }
        ]
    }
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(leads)} leads to {OUT}")


if __name__ == "__main__":
    main()
