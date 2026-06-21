from pathlib import Path
import json

from openpyxl import load_workbook


ROOT = Path("/Users/rishav/Downloads/codex")
XLSX = ROOT / "outputs" / "us_linkedin_startup_leads" / "us_linkedin_startup_leads_40.xlsx"
OUT = ROOT / "outputs" / "us_linkedin_startup_leads" / "us_linkedin_startup_leads_40.crm.json"


def cell_value(value):
    if value is None:
        return ""
    return str(value).strip()


wb = load_workbook(XLSX, data_only=True)
ws = wb["US LinkedIn Leads"]
headers = [cell_value(cell.value) for cell in ws[1]]

records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    record = {headers[index]: cell_value(value) for index, value in enumerate(row) if index < len(headers)}
    if record.get("Company"):
        records.append(record)

if len(records) != 40:
    raise SystemExit(f"Expected 40 leads, found {len(records)}")

OUT.write_text(json.dumps(records, indent=2), encoding="utf-8")
print(OUT)
